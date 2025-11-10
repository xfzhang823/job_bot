# job_bot/tools/underlines_to_selection.py
"""
underlines_to_selection (openpyxl-only)
=======================================

Extract the *human-selected* (underlined) requirement match from an
Alignment Review crosstab into copy/paste–ready CSV and pipeline-friendly JSON.

This version uses **openpyxl only** (pure Python; no Excel/COM dependency).

Workflow
--------
1) Your FSM exports the Alignment Review crosstab XLSX to `EXCEL_DIR`.
2) You open the workbook and **underline exactly one** cell per responsibility row
   (the chosen match).
3) Run this tool. It scans the sheet for underlined cells and emits:
   • CSV  → `__underlined_selection.csv`
   • JSON → `__underlined_selection.json` (mapping: responsibility_key → selected_text)

Configured directories come from `pipeline_config.py`:
- EXCEL_DIR: input workbooks directory
- SEL_DIR  : base directory for extracted selections (per-slug subfolders)

Usage
-----
Run from repo root:

    python -m job_bot.tools.underlines_to_selection \
      "mediabrands_director-intelligence-solutions_4800265007__alignment_review.xlsx" \
      --sheet-name review_grid \
      --log-level INFO

If the given path doesn’t exist as provided, the tool also checks inside EXCEL_DIR.

Outputs
-------
SEL_DIR/<slug>/__underlined_selection.csv
SEL_DIR/<slug>/__underlined_selection.json

Notes
-----
- The tool ignores columns whose headers end with "(Score)" and the column
  named exactly "Original Responsibility".
- If multiple cells are underlined in a row, the first is taken (a WARNING is logged).
- If a row has no underlined cell, it is skipped (DEBUG logs will note counts).

Dependencies
------------
pip install openpyxl pandas
"""

from __future__ import annotations

import argparse
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

# Project paths (centralized in your config)
from job_bot.config.project_config import (
    EXCEL_DIR,
    SEL_DIR,
    TEMP_DIR,
)  # noqa: F401  (TEMP_DIR reserved)

# -----------------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# openpyxl underline values can be: True, "single", "double", "singleAccounting", "doubleAccounting", or None
VALID_UNDERLINES = {True, "single", "double", "singleAccounting", "doubleAccounting"}


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _slugify(s: str) -> str:
    """
    Normalize a string into a filesystem-friendly slug.

    - lowercases
    - drops protocol prefix if it's a URL
    - replaces non-alphanumeric with underscores
    - trims to 200 chars
    """
    s = re.sub(r"^https?://", "", s.strip().lower())
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")[:200]


def _slug_from_xlsx(xlsx_path: Path) -> str:
    """
    Derive a stable slug from a crosstab filename.

    Strips a conventional suffix like '__alignment_review' and optional '__iterN'.

    Examples
    --------
    'myrole_123__alignment_review.xlsx'         → 'myrole_123'
    'myrole_123__alignment_review__iter2.xlsx'  → 'myrole_123'
    """
    stem = xlsx_path.stem
    stem = re.sub(r"__alignment_review(__iter\d+)?$", "", stem, flags=re.IGNORECASE)
    return _slugify(stem)


def _resolve_input_path(xlsx_path: str | Path) -> Path:
    """
    Resolve an input path, checking EXCEL_DIR if the given path is missing.

    Parameters
    ----------
    xlsx_path : str | Path
        Provided path or filename.

    Returns
    -------
    Path
        Resolved existing path to the XLSX file.

    Raises
    ------
    FileNotFoundError
        If not found at the given path nor under EXCEL_DIR.
    """
    p = Path(xlsx_path)
    if p.exists():
        return p
    candidate = EXCEL_DIR / p
    if candidate.exists():
        logger.debug("Resolved input under EXCEL_DIR: %s", candidate)
        return candidate
    raise FileNotFoundError(f"Could not find Excel file at '{p}' or '{candidate}'")


def _prepare_output_dir(slug: str, out_dir: str | Path | None) -> Path:
    """
    Resolve output directory and ensure it exists.

    Parameters
    ----------
    slug : str
        Stable identifier derived from the workbook filename.
    out_dir : str | Path | None
        Optional explicit output directory. If None, defaults to SEL_DIR/<slug>/.

    Returns
    -------
    Path
        Existing output directory.
    """
    final = Path(out_dir) if out_dir is not None else (SEL_DIR / slug)
    final.mkdir(parents=True, exist_ok=True)
    return final


# -----------------------------------------------------------------------------
# Core extractor (openpyxl)
# -----------------------------------------------------------------------------
def extract_underlined_from_xlsx(
    xlsx_path: str | Path,
    *,
    sheet_name: str = "review_grid",
    out_dir: str | Path | None = None,
) -> tuple[pd.DataFrame, Path, Path]:
    """
    Extract underlined selections from an Alignment Review crosstab (openpyxl-only).

    This function scans a reviewed Alignment Review Excel workbook for **underlined**
    requirement cells (human-marked selections) corresponding to each responsibility row.

    It supports both legacy and current sheet names ("review_grid" or "Alignment Review").

    Behavior
    --------
    • Each row in the sheet corresponds to a resume responsibility.
    • Each column (beyond the first two) corresponds to a job requirement.
    • A human reviewer underlines exactly one requirement per responsibility.
    • The function identifies those underlined cells and emits structured output.

    Rules
    -----
    - If **multiple cells are underlined** in one row, the **first** is taken
      (logged with a WARNING).
    - If a row has **no underlined cell**, it is skipped.
      A summary WARNING is logged with the count and sample `responsibility_key`s,
      since this may indicate human error.
    - Columns ending with "(Score)" and the "Original Responsibility" column
      are ignored during scanning.
    - Both "review_grid" and "Alignment Review" sheet names are auto-detected.

    Parameters
    ----------
    xlsx_path : str | Path
        Path or filename of the crosstab workbook.
        If the given path doesn't exist, the function also checks inside `EXCEL_DIR`.
    sheet_name : str, default 'review_grid'
        Optional hint; auto-detection still applies for common names.
    out_dir : str | Path | None
        Optional explicit output directory. If None, writes to `SEL_DIR/<slug>/`.

    Returns
    -------
    df_out : pandas.DataFrame
        Table with columns:
            - responsibility_key : Resume responsibility identifier
            - selected_text      : Underlined requirement text (the human selection)
            - column_header      : Requirement header corresponding to that selection
    csv_path : Path
        Path to the CSV file written.
    json_path : Path
        Path to the JSON file written (mapping responsibility_key → selected_text).

    Warnings
    --------
    - Rows with multiple underlines are logged individually.
    - Rows with no underlines are summarized at the end (with sample keys).

    Example
    -------
    >>> df, csv_path, json_path = extract_underlined_from_xlsx(
    ...     "mediabrands_director-intelligence-solutions_4800265007__alignment_review.xlsx"
    ... )
    >>> df.head()
      responsibility_key                                       selected_text                                     column_header
    0         10003.2.1  Ability to leverage advanced analytics, data science...  Ability to leverage advanced analytics, data science...
    """
    xlsx = _resolve_input_path(xlsx_path)
    slug = _slug_from_xlsx(xlsx)
    out_dir_final = _prepare_output_dir(slug, out_dir)

    logger.info("Starting underline extraction (openpyxl)")
    logger.info("  Input workbook : %s", xlsx)

    wb = load_workbook(filename=xlsx, data_only=True)

    # --- Auto-detect sheet name (legacy/new) ---
    possible_names = ["review_grid", "Alignment Review"]
    for name in possible_names:
        if name in wb.sheetnames:
            sheet_name = name
            logger.info("  Detected sheet name: %s", sheet_name)
            break
    else:
        raise ValueError(f"No expected sheet found. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    logger.info("  Output folder  : %s", out_dir_final)

    # Read headers from the first row
    headers: List[str] = [str(c.value or "").strip() for c in ws[1]]
    ncols = len(headers)
    logger.info("  Header columns : %d", ncols)
    logger.debug("  Headers        : %s", headers)

    # Identify columns to inspect (exclude "(Score)" and "Original Responsibility")
    score_cols = {i for i, h in enumerate(headers, start=1) if h.endswith("(Score)")}
    skip_cols = {
        i for i, h in enumerate(headers, start=1) if h in {"Original Responsibility"}
    }
    text_cols = [i for i in range(1, ncols + 1) if i not in score_cols | skip_cols]

    logger.debug("  Score col idx  : %s", sorted(score_cols))
    logger.debug("  Skip col idx   : %s", sorted(skip_cols))
    logger.debug("  Text col idx   : %s", sorted(text_cols))

    # Assume first column holds the responsibility key label from your export.
    resp_key_col = 1

    rows_out: List[Dict[str, str]] = []
    rows_no_underline = 0
    rows_multi_underline = 0
    no_sel_keys: List[str] = []  # Track missing selections for summary warning

    # Iterate data rows (start at row 2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ncols):
        cells = list(row)
        resp_cell = cells[resp_key_col - 1]
        resp_key = resp_cell.value
        if resp_key is None:
            continue
        resp_key = str(resp_key).strip()

        underlined_cells: List[Tuple[str, str]] = []
        for j in text_cols:
            cell = cells[j - 1]
            val = cell.value
            if val is None:
                continue
            underline = getattr(getattr(cell, "font", None), "underline", None)
            if underline in VALID_UNDERLINES:
                header = headers[j - 1]
                underlined_cells.append((header, str(val).strip()))

        if not underlined_cells:
            rows_no_underline += 1
            no_sel_keys.append(resp_key)
            logger.debug("  No underline on row (resp_key=%s)", resp_key)
            continue

        if len(underlined_cells) > 1:
            rows_multi_underline += 1
            # By design: take the FIRST underlined cell as the chosen match
            logger.warning(
                "Multiple underlines on row (resp_key=%s); taking the first.",
                resp_key,
            )

        header, text = underlined_cells[0]
        rows_out.append(
            {
                "responsibility_key": resp_key,
                "selected_text": text,
                "column_header": header,
            }
        )

    df_out = (
        pd.DataFrame(rows_out).sort_values("responsibility_key").reset_index(drop=True)
    )

    csv_path = out_dir_final / "__underlined_selection.csv"
    json_path = out_dir_final / "__underlined_selection.json"

    df_out.to_csv(csv_path, index=False)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                row["responsibility_key"]: row["selected_text"]
                for _, row in df_out.iterrows()
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    # --- Summary logging for potential human misses ---
    if rows_no_underline > 0:
        sample = ", ".join(no_sel_keys[:10])
        logger.warning(
            "⚠️  %d row(s) have no underlined selection. This may indicate human error.\n"
            "   Sample responsibility_keys: %s%s",
            rows_no_underline,
            sample,
            " ..." if len(no_sel_keys) > 10 else "",
        )
    if rows_multi_underline > 0:
        logger.warning("⚠️  Rows with multiple underlines: %d", rows_multi_underline)

    logger.info("Extraction complete")
    logger.info("  Rows extracted : %s", len(df_out))
    logger.info("  CSV  → %s", csv_path)
    logger.info("  JSON → %s", json_path)

    return df_out, csv_path, json_path


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
def _build_arg_parser() -> argparse.ArgumentParser:
    """
    Create the CLI argument parser.

    Returns
    -------
    argparse.ArgumentParser
        Configured parser for main entrypoint.
    """
    p = argparse.ArgumentParser(
        prog="python -m job_bot.tools.underlines_to_selection",
        description=(
            "Extract underlined selections from an Alignment Review crosstab (openpyxl-only).\n"
            "If the input path is not found, the tool will also search under EXCEL_DIR."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "xlsx",
        help="Path to crosstab workbook, or filename inside EXCEL_DIR.",
    )
    p.add_argument(
        "--sheet-name",
        default="review_grid",
        help="Worksheet name containing the crosstab.",
    )
    p.add_argument(
        "--out-dir",
        default=None,
        help="Optional explicit output directory. If omitted, writes to SEL_DIR/<slug>/.",
    )
    p.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging verbosity.",
    )
    return p


def _configure_root_logger(level: str) -> None:
    """
    Configure the root logger with a simple console handler.

    Parameters
    ----------
    level : str
        One of 'DEBUG', 'INFO', 'WARNING', 'ERROR'.
    """
    logging.basicConfig(
        level=getattr(logging, level.upper(), logging.INFO),
        format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    )


def main() -> int:
    """
    CLI entrypoint.

    Returns
    -------
    int
        Process exit code (0 for success, non-zero for failure).
    """
    parser = _build_arg_parser()
    args = parser.parse_args()

    _configure_root_logger(args.log_level)

    try:
        df, csv_p, json_p = extract_underlined_from_xlsx(
            args.xlsx, sheet_name=args.sheet_name, out_dir=args.out_dir
        )
        print(f"Wrote:\n  CSV : {csv_p}\n  JSON: {json_p}\nRows: {len(df)}")
        return 0
    except Exception as e:
        logger.exception("Extraction failed: %s", e)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
